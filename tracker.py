#Utility script to extract metadata

from pathlib import Path
from openpyxl import Workbook, load_workbook
import os
import sys

EXCEL_FILE = "problems.xlsx"


def parse_metadata(filepath):
    metadata = {
        "problem": "",
        "platform": "",
        "topic": "",
        "description": ""
    }

    with open(filepath, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()

            if not line.startswith("//"):
                break

            text = line[2:].strip()

            if text.startswith("Problem:"):
                metadata["problem"] = text.replace(
                    "Problem:", "", 1
                ).strip()

            elif text.startswith("Platform:"):
                metadata["platform"] = text.replace(
                    "Platform:", "", 1
                ).strip()

            elif text.startswith("Topic:"):
                metadata["topic"] = text.replace(
                    "Topic:", "", 1
                ).strip()

            elif text.startswith("Description:"):
                metadata["description"] = text.replace(
                    "Description:", "", 1
                ).strip()

    return metadata


def load_sheet():

    if Path(EXCEL_FILE).exists():

        wb = load_workbook(EXCEL_FILE)
        ws = wb.active

    else:

        wb = Workbook()
        ws = wb.active

        ws.title = "Problems"

        ws.append([
            "Problem ID",
            "File Name",
            "Platform",
            "Topic",
            "Description",
            "Solved Date",
            "Last Updated"
        ])

        wb.save(EXCEL_FILE)

    return wb, ws


def update_problem(filepath):

    wb, ws = load_sheet()

    metadata = parse_metadata(filepath)

    problem_id = metadata["problem"]

    if not problem_id:
        print(f"Skipping {filepath}: Problem ID missing")
        return

    file_name = Path(filepath).stem

    platform = metadata["platform"]
    topic = metadata["topic"]
    description = metadata["description"]

    push_date = os.environ.get("PUSH_DATE")

    if push_date:
        today = push_date[:10]
    else:
        from datetime import datetime
        today = datetime.utcnow().strftime("%Y-%m-%d")

    existing_row = None

    for row in range(2, ws.max_row + 1):

        current_problem_id = str(
            ws.cell(row=row, column=1).value
        )

        if current_problem_id == str(problem_id):
            existing_row = row
            break

    if existing_row:

        ws.cell(existing_row, 2).value = file_name
        ws.cell(existing_row, 3).value = platform
        ws.cell(existing_row, 4).value = topic
        ws.cell(existing_row, 5).value = description

        # Update only Last Updated
        ws.cell(existing_row, 7).value = today

        print(
            f"Updated Problem {problem_id} ({file_name})"
        )

    else:

        ws.append([
            problem_id,
            file_name,
            platform,
            topic,
            description,
            today,
            today
        ])

        print(
            f"Added Problem {problem_id} ({file_name})"
        )

    wb.save(EXCEL_FILE)


if __name__ == "__main__":

    if len(sys.argv) != 2:
        print("Usage: python tracker.py <cpp_file>")
        sys.exit(1)

    update_problem(sys.argv[1])